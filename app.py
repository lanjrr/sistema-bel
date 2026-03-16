import streamlit as st
import psycopg2
import psycopg2.extras
import pandas as pd
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ══════════════════════════════════════════════
# BANCO DE DADOS — SUPABASE / POSTGRESQL
# ══════════════════════════════════════════════
def get_conn():
    url = st.secrets["DATABASE_URL"]
    return psycopg2.connect(url, sslmode="require", connect_timeout=10)

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS modelos
                 (id SERIAL PRIMARY KEY, nome TEXT UNIQUE)''')
    c.execute('''CREATE TABLE IF NOT EXISTS clientes
                 (id SERIAL PRIMARY KEY, nome TEXT UNIQUE)''')
    c.execute('''CREATE TABLE IF NOT EXISTS producao
                 (id SERIAL PRIMARY KEY,
                  di TEXT, modelo TEXT, modelo_original TEXT, cliente TEXT,
                  serial_china TEXT UNIQUE, serial_brasil TEXT,
                  pedido TEXT, valor_antes TEXT, valor_depois TEXT,
                  exc_se TEXT, exc_sd TEXT, exc_ie TEXT, exc_id TEXT,
                  carga_maxima TEXT, zero TEXT, numero_lacre TEXT,
                  status TEXT)''')
    conn.commit()
    conn.close()

init_db()

def query_df(sql, params=None):
    conn = get_conn()
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df

def execute(sql, params=None):
    conn = get_conn()
    c = conn.cursor()
    c.execute(sql, params or ())
    conn.commit()
    conn.close()

def executemany_safe(rows_sql, rows_data, serial_idx=3):
    conn = get_conn()
    c = conn.cursor()
    ok, dup = 0, []
    for data in rows_data:
        try:
            c.execute(rows_sql, data)
            conn.commit()
            ok += 1
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            dup.append(data[serial_idx])
    conn.close()
    return ok, dup

def parsear_seriais(texto: str) -> list:
    """
    Aceita seriais separados por newline, espaço, vírgula ou tabulação.
    Retorna lista de seriais únicos não vazios.
    """
    import re
    tokens = re.split(r'[\n\r,;\t ]+', texto.strip())
    return [t.strip() for t in tokens if t.strip()]

# ══════════════════════════════════════════════
# GERAÇÃO DE PDF
# ══════════════════════════════════════════════
def gerar_pdf_relatorio(row) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    cor_bel = colors.HexColor('#1a1a2e')
    cor_azul = colors.HexColor('#0066cc')

    estilo_titulo = ParagraphStyle('titulo', fontSize=18, textColor=cor_bel,
                                   fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    estilo_sub = ParagraphStyle('sub', fontSize=10, textColor=colors.grey,
                                alignment=TA_CENTER, spaceAfter=2)
    estilo_secao = ParagraphStyle('secao', fontSize=11, textColor=cor_azul,
                                  fontName='Helvetica-Bold', spaceBefore=14, spaceAfter=6)
    estilo_normal = ParagraphStyle('normal', fontSize=10, textColor=cor_bel)

    elementos = []

    # Cabeçalho
    elementos.append(Paragraph("BEL EQUIPAMENTOS ANALÍTICOS LTDA", estilo_titulo))
    elementos.append(Paragraph("Rua Alferes José Caetano, 1572 — Piracicaba / SP", estilo_sub))
    elementos.append(Paragraph("INMETRO/DIMEL nº 0008/2018", estilo_sub))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(HRFlowable(width="100%", thickness=2, color=cor_bel))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph("RELATÓRIO DE AFERIÇÃO", estilo_titulo))
    elementos.append(Spacer(1, 0.4*cm))

    # Identificação da balança
    elementos.append(Paragraph("Identificação do Equipamento", estilo_secao))
    dados_id = [
        ["Serial Brasil (Inmetro)", row.get('serial_brasil') or '—',
         "Serial China (Fábrica)", row.get('serial_china') or '—'],
        ["Modelo", row.get('modelo') or '—',
         "Modelo Original", row.get('modelo_original') or row.get('modelo') or '—'],
        ["DI", row.get('di') or '—',
         "Nº do Lacre", row.get('numero_lacre') or '—'],
    ]
    t_id = Table(dados_id, colWidths=[4.5*cm, 5.5*cm, 4.5*cm, 5.5*cm])
    t_id.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f0f4ff')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#f0f4ff')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUND', (0,0), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elementos.append(t_id)

    # Destino comercial
    elementos.append(Paragraph("Destino Comercial", estilo_secao))
    dados_com = [
        ["Cliente", row.get('cliente') or '—', "Pedido", row.get('pedido') or '—'],
    ]
    t_com = Table(dados_com, colWidths=[4.5*cm, 5.5*cm, 4.5*cm, 5.5*cm])
    t_com.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f0f4ff')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#f0f4ff')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elementos.append(t_com)

    # Dados de calibração
    elementos.append(Paragraph("Dados de Calibração", estilo_secao))
    dados_cal = [
        ["Indicação Antes", "Indicação Depois", "Carga Máxima", "Zero"],
        [row.get('valor_antes') or '—', row.get('valor_depois') or '—',
         row.get('carga_maxima') or '—', row.get('zero') or '—'],
    ]
    t_cal = Table(dados_cal, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
    t_cal.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), cor_bel),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#f8f9fa')),
    ]))
    elementos.append(t_cal)

    # Excentricidade
    elementos.append(Paragraph("Teste de Excentricidade (4 cantos)", estilo_secao))
    dados_exc = [
        ["Sup. Esquerdo", "Sup. Direito", "Inf. Esquerdo", "Inf. Direito"],
        [row.get('exc_se') or '—', row.get('exc_sd') or '—',
         row.get('exc_ie') or '—', row.get('exc_id') or '—'],
    ]
    t_exc = Table(dados_exc, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
    t_exc.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2d4060')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#f8f9fa')),
    ]))
    elementos.append(t_exc)

    # Rodapé
    elementos.append(Spacer(1, 1*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cccccc')))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(
        "Documento gerado pelo Sistema de Rastreabilidade — Bel Equipamentos Analíticos Ltda",
        ParagraphStyle('rodape', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(elementos)
    return buf.getvalue()


# ══════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════
def estilizar_excel(ws, df):
    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)),
                      df[col].astype(str).map(len).max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 40)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.alignment = Alignment(horizontal="center")

def gerar_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
        estilizar_excel(writer.sheets['Dados'], df)
    return buf.getvalue()

def gerar_planilha_afericao(df_sel: pd.DataFrame) -> bytes:
    colunas_vazias = [
        "Novo_Serial_Brasil", "Numero_Lacre",
        "Valor_Antes", "Valor_Depois",
        "Exc_Sup_Esq", "Exc_Sup_Dir", "Exc_Inf_Esq", "Exc_Inf_Dir",
        "Carga_Max", "Zero"
    ]
    df_out = df_sel[["Serial_China", "Modelo", "DI"]].copy()
    for col in colunas_vazias:
        df_out[col] = ""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='Calibracao')
        ws = writer.sheets['Calibracao']
        estilizar_excel(ws, df_out)
        fill_pre = PatternFill("solid", fgColor="dce8f5")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill_pre
    return buf.getvalue()


# ══════════════════════════════════════════════
# CONFIGURAÇÃO GERAL
# ══════════════════════════════════════════════
st.set_page_config(page_title="Bel Traceability", layout="wide")

st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #1a1a2e; }
    [data-testid="stSidebar"] * { color: #ffffff !important; }
    [data-testid="stSidebar"] .stRadio > label {
        color: #cccccc !important; font-size: 11px;
        text-transform: uppercase; letter-spacing: 1px;
    }
    h1 { color: #ffffff !important; }
    h2 { color: #ffffff !important; border-bottom: 2px solid #333; padding-bottom: 8px; }
    h3 { color: #ffffff !important; }
    [data-testid="metric-container"] {
        background: #1e2a3a; border: 1px solid #2d4060;
        border-radius: 8px; padding: 12px;
    }
    .info-box {
        background: #1e2a3a; border-left: 4px solid #0066cc;
        border-radius: 4px; padding: 12px 16px; margin: 8px 0;
        font-size: 14px; color: #ffffff !important;
    }
    .secao-titulo {
        font-size: 17px; font-weight: 600; color: #ffffff;
        margin-top: 8px; margin-bottom: 2px;
    }
    .detalhe-serial-br { font-size: 26px; color: #4da6ff; font-weight: 700; }
    .detalhe-serial-china { font-size: 14px; color: #aaaaaa; margin-top: 4px; }
    .detalhe-label { font-size: 11px; color: #aaaaaa; text-transform: uppercase; letter-spacing: 1px; }
    .badge-disponivel  { background:#0066cc; color:#fff; padding:4px 12px; border-radius:12px; font-size:13px; }
    .badge-pronta      { background:#7b2d8b; color:#fff; padding:4px 12px; border-radius:12px; font-size:13px; }
    .badge-finalizado  { background:#007a33; color:#fff; padding:4px 12px; border-radius:12px; font-size:13px; }
    /* Grade de seriais */
    .grade-serial {
        display: flex; flex-wrap: wrap; gap: 6px; margin: 8px 0;
    }
    .serial-cell {
        background: #1e2a3a; border: 1px solid #2d4060; border-radius: 4px;
        padding: 4px 10px; font-size: 12px; color: #ffffff;
        font-family: monospace;
    }
</style>
""", unsafe_allow_html=True)


def color_status(val):
    if val == 'Disponível':     return 'color: #4da6ff; font-weight: bold'
    if val == 'Finalizado':     return 'color: #00cc66; font-weight: bold'
    if val == 'Pronta Entrega': return 'color: #cc88ff; font-weight: bold'
    return 'color: #ffaa44; font-weight: bold'

def exibir_grade_seriais(seriais: list):
    """Exibe lista de seriais em formato de grade visual."""
    cells = "".join(f'<span class="serial-cell">{s}</span>' for s in seriais)
    st.markdown(f'<div class="grade-serial">{cells}</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════
# MENU LATERAL
# ══════════════════════════════════════════════
st.sidebar.title("🌐 Bel Engineering")
st.sidebar.markdown("**Sistema de Rastreabilidade**")
st.sidebar.markdown("---")

unidade = st.sidebar.selectbox(
    "Unidade Operacional", ["Brasil", "China (Factory)"],
    index=None, placeholder="Selecione a unidade..."
)

# ══════════════════════════════════════════════
# BRASIL
# ══════════════════════════════════════════════
if unidade == "Brasil":
    st.sidebar.markdown("---")
    menu = st.sidebar.radio("Navegação", [
        "📊 Dashboard",
        "📋 Gestão de Cadastros",
        "🔬 Bancada de Aferição",
        "📦 Pronta Entrega",
        "🔍 Consulta",
        "⚙️  Configurações",
    ])

    # ══════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════
    if menu == "📊 Dashboard":
        st.title("📊 Dashboard — Visão Geral")

        df_all = query_df("SELECT status, modelo, di FROM producao")
        total  = len(df_all)
        disp   = len(df_all[df_all['status'] == 'Disponível'])
        pronta = len(df_all[df_all['status'] == 'Pronta Entrega'])
        final  = len(df_all[df_all['status'] == 'Finalizado'])

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total de Balanças", total)
        c2.metric("🔵 Disponíveis", disp)
        c3.metric("🟣 Pronta Entrega", pronta)
        c4.metric("🟢 Finalizadas", final)

        if not df_all.empty:
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Por Status")
                df_s = df_all['status'].value_counts().reset_index()
                df_s.columns = ['Status', 'Quantidade']
                st.dataframe(df_s, use_container_width=True, hide_index=True)
            with col2:
                st.subheader("Por Modelo")
                df_m = df_all['modelo'].value_counts().reset_index()
                df_m.columns = ['Modelo', 'Quantidade']
                st.dataframe(df_m, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("Por DI")
            df_di = query_df("""
                SELECT di AS "DI", COUNT(*) AS "Total",
                       SUM(CASE WHEN status='Disponível'     THEN 1 ELSE 0 END) AS "Disponíveis",
                       SUM(CASE WHEN status='Pronta Entrega' THEN 1 ELSE 0 END) AS "Pronta Entrega",
                       SUM(CASE WHEN status='Finalizado'     THEN 1 ELSE 0 END) AS "Finalizadas",
                       STRING_AGG(DISTINCT modelo, ', ') AS "Modelos"
                FROM producao WHERE di IS NOT NULL AND di != ''
                GROUP BY di ORDER BY di DESC
            """)
            if not df_di.empty:
                st.dataframe(df_di, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum dado registrado ainda.")

    # ══════════════════════════════════════════
    # GESTÃO DE CADASTROS
    # ══════════════════════════════════════════
    elif menu == "📋 Gestão de Cadastros":
        st.title("📋 Gestão de Cadastros")

        tab_modelos, tab_clientes, tab_dis = st.tabs([
            "🏷️  Modelos", "👥 Clientes", "🚢 DIs / Recebimento"
        ])

        # ── MODELOS ───────────────────────────
        with tab_modelos:
            df_mod = query_df("SELECT id, nome FROM modelos ORDER BY nome")

            st.markdown('<p class="secao-titulo">➕ Novo Modelo</p>', unsafe_allow_html=True)
            with st.form("form_novo_modelo", clear_on_submit=True):
                ci, cb = st.columns([3, 1])
                novo_mod = ci.text_input("m", label_visibility="collapsed",
                                          placeholder="Nome do modelo (ex: M214Ai)").strip()
                if cb.form_submit_button("➕ Inserir", use_container_width=True) and novo_mod:
                    try:
                        execute("INSERT INTO modelos (nome) VALUES (%s)", (novo_mod,))
                        st.success(f"Modelo '{novo_mod}' cadastrado!")
                        st.rerun()
                    except Exception:
                        st.error("Modelo já existe.")

            st.markdown("---")
            st.markdown('<p class="secao-titulo">📋 Modelos Cadastrados</p>', unsafe_allow_html=True)

            if df_mod.empty:
                st.info("Nenhum modelo cadastrado ainda.")
            else:
                cl, ca = st.columns([2, 1])
                with cl:
                    st.dataframe(df_mod[['nome']].rename(columns={'nome': 'Modelo'}),
                                 use_container_width=True, hide_index=True)
                with ca:
                    st.markdown("**Editar / Excluir**")
                    mid = st.selectbox("sel", df_mod['id'].tolist(),
                        format_func=lambda i: df_mod[df_mod['id']==i]['nome'].values[0],
                        index=None, placeholder="Escolha...", key="sel_mod",
                        label_visibility="collapsed")
                    if mid:
                        nm = df_mod[df_mod['id']==mid]['nome'].values[0]
                        with st.form("form_edit_mod"):
                            nn = st.text_input("Nome", value=nm)
                            b1, b2 = st.columns(2)
                            if b1.form_submit_button("💾 Salvar", use_container_width=True):
                                try:
                                    execute("UPDATE modelos SET nome=%s WHERE id=%s", (nn.strip(), mid))
                                    st.success("Atualizado!")
                                    st.rerun()
                                except Exception:
                                    st.error("Nome já existe.")
                            if b2.form_submit_button("🗑️ Excluir", use_container_width=True):
                                em_uso = query_df(
                                    "SELECT COUNT(*) as n FROM producao WHERE modelo=%s", params=(nm,)
                                )['n'].values[0]
                                if em_uso:
                                    st.error(f"Em uso por {em_uso} balança(s).")
                                else:
                                    execute("DELETE FROM modelos WHERE id=%s", (mid,))
                                    st.success("Excluído.")
                                    st.rerun()

        # ── CLIENTES ──────────────────────────
        with tab_clientes:
            df_cli     = query_df("SELECT id, nome FROM clientes ORDER BY nome")
            df_cli_cnt = query_df(
                "SELECT cliente, COUNT(*) as total FROM producao WHERE cliente IS NOT NULL AND cliente!='' GROUP BY cliente")

            st.markdown('<p class="secao-titulo">➕ Novo Cliente</p>', unsafe_allow_html=True)
            with st.form("form_novo_cli", clear_on_submit=True):
                ci, cb = st.columns([3, 1])
                novo_cli = ci.text_input("c", label_visibility="collapsed",
                                          placeholder="Nome do cliente").strip()
                if cb.form_submit_button("➕ Inserir", use_container_width=True) and novo_cli:
                    try:
                        execute("INSERT INTO clientes (nome) VALUES (%s)", (novo_cli,))
                        st.success(f"Cliente '{novo_cli}' cadastrado!")
                        st.rerun()
                    except Exception:
                        st.error("Cliente já existe.")

            st.markdown("---")
            st.markdown('<p class="secao-titulo">📋 Clientes Cadastrados</p>', unsafe_allow_html=True)

            if df_cli.empty:
                st.info("Nenhum cliente cadastrado ainda.")
            else:
                cl, ca = st.columns([2, 1])
                with cl:
                    df_d = df_cli.merge(df_cli_cnt, left_on='nome', right_on='cliente', how='left')
                    df_d['total'] = df_d['total'].fillna(0).astype(int)
                    st.dataframe(df_d[['nome','total']].rename(
                        columns={'nome':'Cliente','total':'Balanças vinculadas'}),
                        use_container_width=True, hide_index=True)
                with ca:
                    st.markdown("**Editar / Excluir**")
                    cid = st.selectbox("sel", df_cli['id'].tolist(),
                        format_func=lambda i: df_cli[df_cli['id']==i]['nome'].values[0],
                        index=None, placeholder="Escolha...", key="sel_cli",
                        label_visibility="collapsed")
                    if cid:
                        nc = df_cli[df_cli['id']==cid]['nome'].values[0]
                        with st.form("form_edit_cli"):
                            nn = st.text_input("Nome", value=nc)
                            b1, b2 = st.columns(2)
                            if b1.form_submit_button("💾 Salvar", use_container_width=True):
                                try:
                                    execute("UPDATE clientes SET nome=%s WHERE id=%s", (nn.strip(), cid))
                                    execute("UPDATE producao SET cliente=%s WHERE cliente=%s", (nn.strip(), nc))
                                    st.success("Atualizado!")
                                    st.rerun()
                                except Exception:
                                    st.error("Nome já existe.")
                            if b2.form_submit_button("🗑️ Excluir", use_container_width=True):
                                em_uso = query_df(
                                    "SELECT COUNT(*) as n FROM producao WHERE cliente=%s", params=(nc,)
                                )['n'].values[0]
                                if em_uso:
                                    st.error(f"Em uso por {em_uso} balança(s).")
                                else:
                                    execute("DELETE FROM clientes WHERE id=%s", (cid,))
                                    st.success("Excluído.")
                                    st.rerun()

        # ── DIs / RECEBIMENTO ─────────────────
        with tab_dis:
            lista_modelos = query_df("SELECT nome FROM modelos ORDER BY nome")['nome'].tolist()

            st.markdown('<p class="secao-titulo">➕ Registrar Nova Entrada de Equipamentos</p>',
                        unsafe_allow_html=True)

            if not lista_modelos:
                st.warning("Cadastre pelo menos um modelo na aba **🏷️ Modelos** antes de registrar uma DI.")
            else:
                di_numero = st.text_input("Número da DI (ex: DI26034)", key="di_input")
                st.markdown("---")
                st.markdown("**Modelos e Seriais desta DI**")
                st.markdown(
                    '<div class="info-box">Adicione um bloco por modelo. Os seriais podem ser colados um por linha, separados por espaço, vírgula ou tabulação.</div>',
                    unsafe_allow_html=True)

                if 'num_blocos' not in st.session_state:
                    st.session_state.num_blocos = 1

                if st.button("➕ Adicionar outro modelo"):
                    st.session_state.num_blocos += 1
                    st.rerun()

                blocos = []
                for i in range(st.session_state.num_blocos):
                    st.markdown(f"**Modelo {i+1}**")
                    cm_col, cr_col = st.columns([4, 1])
                    with cm_col:
                        mod_b = st.selectbox("Modelo", lista_modelos,
                            index=None, placeholder="Selecione o modelo...",
                            key=f"mod_bloco_{i}")
                    with cr_col:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.session_state.num_blocos > 1:
                            if st.button("🗑️", key=f"rem_{i}"):
                                st.session_state.num_blocos -= 1
                                st.rerun()
                    sn_b = st.text_area("Seriais (um por linha, ou separados por espaço/vírgula)",
                        height=110, key=f"sn_bloco_{i}")

                    # Preview em grade
                    if sn_b and sn_b.strip():
                        sn_preview = parsear_seriais(sn_b)
                        if sn_preview:
                            st.markdown(f"**{len(sn_preview)} serial(is) detectado(s):**")
                            exibir_grade_seriais(sn_preview)

                    blocos.append((mod_b, sn_b))
                    st.markdown("---")

                if st.button("📥 Dar Entrada no Estoque", type="primary", use_container_width=True):
                    di_v    = st.session_state.get("di_input", "").strip()
                    validos = [(m, s) for m, s in blocos if m and s and s.strip()]

                    if not validos:
                        st.error("Preencha ao menos um modelo com seriais.")
                    else:
                        rows = []
                        for mod_v, sn_v in validos:
                            for sn in parsear_seriais(sn_v):
                                rows.append((di_v, mod_v, sn, "Disponível"))

                        ok, dup = executemany_safe(
                            "INSERT INTO producao (di, modelo, serial_china, status) VALUES (%s, %s, %s, %s)",
                            rows, serial_idx=2
                        )
                        if ok:  st.success(f"✅ {ok} equipamento(s) registrado(s) como Disponível.")
                        if dup: st.error(f"🚫 Duplicados ignorados: {', '.join(dup)}")
                        st.session_state.num_blocos = 1
                        st.rerun()

            st.markdown("---")
            st.markdown('<p class="secao-titulo">📋 DIs Registradas</p>', unsafe_allow_html=True)

            df_di_res = query_df("""
                SELECT di AS "DI", COUNT(*) AS "Total",
                       SUM(CASE WHEN status='Disponível'     THEN 1 ELSE 0 END) AS "Disponíveis",
                       SUM(CASE WHEN status='Pronta Entrega' THEN 1 ELSE 0 END) AS "Pronta Entrega",
                       SUM(CASE WHEN status='Finalizado'     THEN 1 ELSE 0 END) AS "Finalizadas",
                       STRING_AGG(DISTINCT modelo, ', ') AS "Modelos"
                FROM producao WHERE di IS NOT NULL AND di != ''
                GROUP BY di ORDER BY di DESC
            """)

            if df_di_res.empty:
                st.info("Nenhuma DI registrada ainda.")
            else:
                st.dataframe(df_di_res, use_container_width=True, hide_index=True)
                st.markdown("---")
                st.markdown('<p class="secao-titulo">🔎 Detalhar uma DI</p>', unsafe_allow_html=True)
                di_esc = st.selectbox("Selecione a DI", df_di_res['DI'].tolist(),
                    index=None, placeholder="Escolha uma DI...")
                if di_esc:
                    df_det = query_df("""
                        SELECT serial_china AS "Serial China", serial_brasil AS "Serial Brasil",
                               modelo AS "Modelo", modelo_original AS "Modelo Original",
                               cliente AS "Cliente", pedido AS "Pedido", status AS "Status"
                        FROM producao WHERE di=%s ORDER BY modelo, serial_china
                    """, params=(di_esc,))
                    st.markdown(f"**{len(df_det)} balança(s) na DI {di_esc}:**")

                    # Grade dos seriais disponíveis
                    sn_disp = df_det[df_det['Status']=='Disponível']['Serial China'].tolist()
                    if sn_disp:
                        st.markdown(f"**Disponíveis ({len(sn_disp)}):**")
                        exibir_grade_seriais(sn_disp)

                    st.dataframe(df_det.style.map(color_status, subset=['Status']),
                                 use_container_width=True, hide_index=True)
                    st.download_button(f"⬇️ Exportar DI {di_esc}",
                        data=gerar_excel(df_det), file_name=f"DI_{di_esc}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ══════════════════════════════════════════
    # BANCADA DE AFERIÇÃO
    # ══════════════════════════════════════════
    elif menu == "🔬 Bancada de Aferição":
        st.title("🔬 Bancada de Calibração")

        lista_clientes = query_df("SELECT nome FROM clientes ORDER BY nome")['nome'].tolist()
        lista_todos_modelos = query_df("SELECT nome FROM modelos ORDER BY nome")['nome'].tolist()

        df_disp = query_df("""
            SELECT id, serial_china AS "Serial_China", modelo AS "Modelo", di AS "DI"
            FROM producao WHERE status='Disponível'
            ORDER BY di, modelo, serial_china
        """)

        if df_disp.empty:
            st.warning("Não há balanças disponíveis no estoque no momento.")
        else:
            # ── PASSO 1: Selecionar balanças ──
            st.subheader("1️⃣ Selecione as Balanças a Aferir")

            dis_disp = ["Todas as DIs"] + sorted(df_disp['DI'].dropna().unique().tolist())
            filtro_di = st.selectbox("Filtrar por DI (opcional):", dis_disp, index=0)

            df_filtrada = df_disp if filtro_di == "Todas as DIs" else df_disp[df_disp['DI'] == filtro_di]

            st.markdown(f"**{len(df_filtrada)} balança(s) disponível(is):**")

            # Grade visual das disponíveis
            exibir_grade_seriais(df_filtrada['Serial_China'].tolist())

            st.markdown("---")
            selecionadas = st.multiselect(
                "Selecione os seriais que pegou do estoque:",
                df_filtrada['Serial_China'].tolist(),
                placeholder="Digite ou clique para selecionar...")

            if selecionadas:
                df_sel = df_filtrada[df_filtrada['Serial_China'].isin(selecionadas)].copy()

                # ── PASSO 2: Transformação de modelo ──
                st.markdown("---")
                st.subheader("2️⃣ Verificar / Transformar Modelos")
                st.markdown(
                    '<div class="info-box">Se alguma balança será aferida com um modelo diferente do importado, selecione o novo modelo abaixo. Deixe em branco para manter o modelo original.</div>',
                    unsafe_allow_html=True)

                transformacoes = {}
                for _, row in df_sel.iterrows():
                    col_sn, col_mod_orig, col_seta, col_mod_novo = st.columns([2, 2, 0.5, 2])
                    col_sn.markdown(f"**`{row['Serial_China']}`**")
                    col_mod_orig.markdown(f"Atual: **{row['Modelo']}**")
                    col_seta.markdown("→")
                    novo_mod = col_mod_novo.selectbox(
                        "Novo modelo", ["(manter original)"] + lista_todos_modelos,
                        index=0, key=f"transf_{row['Serial_China']}")
                    if novo_mod != "(manter original)":
                        transformacoes[row['Serial_China']] = novo_mod

                if transformacoes:
                    st.markdown(f"**{len(transformacoes)} transformação(ões) definida(s):**")
                    for sn, nm in transformacoes.items():
                        mod_orig = df_sel[df_sel['Serial_China']==sn]['Modelo'].values[0]
                        st.markdown(f"- `{sn}`: {mod_orig} → **{nm}**")

                # ── PASSO 3: Destino ──────────
                st.markdown("---")
                st.subheader("3️⃣ Destino das Balanças")
                destino = st.radio("", ["Vincular a Cliente", "Pronta Entrega"],
                                   horizontal=True, label_visibility="collapsed")

                if destino == "Vincular a Cliente":
                    col1, col2 = st.columns(2)
                    pedido_sel  = col1.text_input("Número do Pedido Comercial")
                    cliente_sel = col2.selectbox("Cliente Destino", lista_clientes,
                        index=None, placeholder="Selecione o cliente...")
                    status_destino = "Finalizado"
                else:
                    st.markdown(
                        '<div class="info-box">As balanças serão marcadas como <strong>Pronta Entrega</strong>. '
                        'O cliente e pedido poderão ser vinculados depois na aba 📦 Pronta Entrega.</div>',
                        unsafe_allow_html=True)
                    pedido_sel = cliente_sel = ""
                    status_destino = "Pronta Entrega"

                # ── PASSO 4: Planilha ─────────
                st.markdown("---")
                st.subheader("4️⃣ Baixe a Planilha Pré-Preenchida")

                # Aplica transformações no df para a planilha
                df_planilha = df_sel.copy()
                for sn, nm in transformacoes.items():
                    df_planilha.loc[df_planilha['Serial_China']==sn, 'Modelo'] = nm

                st.markdown(
                    '<div class="info-box">A planilha já vem com <strong>Serial China</strong>, '
                    '<strong>Modelo</strong> (após transformação) e <strong>DI</strong> preenchidos. '
                    'Preencha os dados de calibração, o Serial Brasil e o Número do Lacre na bancada.</div>',
                    unsafe_allow_html=True)
                st.download_button(
                    label=f"⬇️ Baixar Planilha ({len(selecionadas)} balança(s))",
                    data=gerar_planilha_afericao(df_planilha),
                    file_name="Afericao_Bel.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # ── PASSO 5: Upload ───────────
                st.markdown("---")
                st.subheader("5️⃣ Suba a Planilha Preenchida")
                arquivo = st.file_uploader("Arraste o arquivo Excel preenchido aqui", type=['xlsx'])

                if arquivo:
                    df_imp = pd.read_excel(arquivo).dropna(how='all')
                    st.write("Pré-visualização:")
                    st.dataframe(df_imp, use_container_width=True)

                    if st.button("💾 Validar e Salvar", type="primary"):
                        if destino == "Vincular a Cliente" and (not cliente_sel or not pedido_sel):
                            st.error("⚠️ Informe o Cliente e o Pedido antes de salvar.")
                        elif "Serial_China" not in df_imp.columns:
                            st.error("⚠️ Use o template gerado no passo 4.")
                        else:
                            conn = get_conn()
                            cur  = conn.cursor()
                            ok, erros = 0, []
                            for _, row in df_imp.iterrows():
                                sc = str(row["Serial_China"]).strip() if not pd.isna(row.get("Serial_China")) else ""
                                if not sc: continue
                                cur.execute(
                                    "SELECT id, modelo FROM producao WHERE serial_china=%s AND status='Disponível'", (sc,))
                                check = cur.fetchone()
                                if check:
                                    rid, mod_atual = check
                                    def s(col): return "" if pd.isna(row.get(col)) else str(row.get(col)).strip()
                                    # Modelo final pode ter sido transformado
                                    mod_final = s("Modelo") or mod_atual
                                    mod_orig  = mod_atual if mod_final != mod_atual else None
                                    cur.execute('''UPDATE producao SET
                                        cliente=%s, pedido=%s, serial_brasil=%s,
                                        modelo=%s, modelo_original=%s,
                                        valor_antes=%s, valor_depois=%s,
                                        exc_se=%s, exc_sd=%s, exc_ie=%s, exc_id=%s,
                                        carga_maxima=%s, zero=%s, numero_lacre=%s,
                                        status=%s WHERE serial_china=%s''',
                                        (cliente_sel, pedido_sel, s("Novo_Serial_Brasil"),
                                         mod_final, mod_orig,
                                         s("Valor_Antes"), s("Valor_Depois"),
                                         s("Exc_Sup_Esq"), s("Exc_Sup_Dir"),
                                         s("Exc_Inf_Esq"), s("Exc_Inf_Dir"),
                                         s("Carga_Max"), s("Zero"), s("Numero_Lacre"),
                                         status_destino, sc))
                                    ok += 1
                                else:
                                    erros.append(sc)
                            conn.commit(); conn.close()
                            if ok:    st.success(f"✅ {ok} balança(s) salvas com status '{status_destino}'!")
                            if erros: st.warning(f"⚠️ Não encontrados ou já finalizados: {', '.join(erros)}")
            else:
                st.markdown("---")
                st.info("Selecione ao menos uma balança acima para continuar.")

    # ══════════════════════════════════════════
    # PRONTA ENTREGA
    # ══════════════════════════════════════════
    elif menu == "📦 Pronta Entrega":
        st.title("📦 Pronta Entrega — Vincular Cliente")

        lista_clientes = query_df("SELECT nome FROM clientes ORDER BY nome")['nome'].tolist()
        df_pe = query_df("""
            SELECT id, serial_china AS "Serial China", serial_brasil AS "Serial Brasil",
                   modelo AS "Modelo", di AS "DI"
            FROM producao WHERE status='Pronta Entrega'
            ORDER BY modelo, serial_brasil
        """)

        if df_pe.empty:
            st.info("Nenhuma balança em Pronta Entrega no momento.")
        else:
            st.markdown(f"**{len(df_pe)} balança(s) disponível(is) para venda:**")
            exibir_grade_seriais(
                [r['Serial Brasil'] or r['Serial China'] for _, r in df_pe.iterrows()])
            st.dataframe(df_pe.drop(columns=['id']), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.subheader("Vincular ao Cliente")

            opcoes = df_pe.apply(
                lambda r: f"{r['Serial Brasil'] or r['Serial China']}  |  {r['Modelo']}  |  DI: {r['DI'] or '—'}",
                axis=1).tolist()
            ids_map = dict(zip(opcoes, df_pe['id'].tolist()))

            selecionadas = st.multiselect("Selecione as balanças a vincular:",
                opcoes, placeholder="Clique para selecionar...")

            col1, col2 = st.columns(2)
            pedido_pe  = col1.text_input("Número do Pedido Comercial", key="pedido_pe")
            cliente_pe = col2.selectbox("Cliente", lista_clientes,
                index=None, placeholder="Selecione o cliente...", key="cliente_pe")

            if selecionadas:
                st.markdown(f"**{len(selecionadas)} balança(s) selecionada(s)**")

            if st.button("✅ Confirmar Venda", type="primary", disabled=not selecionadas):
                if not cliente_pe or not pedido_pe:
                    st.error("Informe o Cliente e o Número do Pedido.")
                else:
                    conn = get_conn()
                    cur  = conn.cursor()
                    for rid in [ids_map[s] for s in selecionadas]:
                        cur.execute(
                            "UPDATE producao SET cliente=%s, pedido=%s, status='Finalizado' WHERE id=%s",
                            (cliente_pe, pedido_pe, rid))
                    conn.commit(); conn.close()
                    st.success(f"✅ {len(selecionadas)} balança(s) vinculada(s) ao cliente {cliente_pe}!")
                    st.rerun()

    # ══════════════════════════════════════════
    # CONSULTA + DETALHE
    # ══════════════════════════════════════════
    elif menu == "🔍 Consulta":
        st.title("🔍 Consulta e Inventário Global")

        if 'detalhe_id' in st.session_state and st.session_state.detalhe_id:
            # ── TELA DE DETALHE ───────────────
            row_df = query_df("""
                SELECT serial_china, serial_brasil, di, modelo, modelo_original,
                       cliente, pedido, valor_antes, valor_depois,
                       exc_se, exc_sd, exc_ie, exc_id,
                       carga_maxima, zero, numero_lacre, status
                FROM producao WHERE id=%s
            """, params=(st.session_state.detalhe_id,))

            if not row_df.empty:
                r = row_df.iloc[0].to_dict()

                col_btn, col_pdf = st.columns([1, 1])
                if col_btn.button("← Voltar para a Consulta"):
                    st.session_state.detalhe_id = None
                    st.rerun()

                # Botão PDF só para balanças finalizadas ou pronta entrega
                if r['status'] in ('Finalizado', 'Pronta Entrega'):
                    pdf_bytes = gerar_pdf_relatorio(r)
                    serial_nome = r.get('serial_brasil') or r.get('serial_china') or 'relatorio'
                    col_pdf.download_button(
                        "📄 Baixar Relatório PDF",
                        data=pdf_bytes,
                        file_name=f"Relatorio_{serial_nome}.pdf",
                        mime="application/pdf",
                        type="primary"
                    )

                st.markdown("---")
                st.markdown(f"""
                <div style="margin-bottom:20px;">
                    <div class="detalhe-label">Serial Brasil (Inmetro)</div>
                    <div class="detalhe-serial-br">{r.get('serial_brasil') or '—'}</div>
                    <div class="detalhe-serial-china">🇨🇳 Serial China: <strong>{r.get('serial_china') or '—'}</strong></div>
                </div>
                """, unsafe_allow_html=True)

                badge_map = {'Disponível':'badge-disponivel','Pronta Entrega':'badge-pronta','Finalizado':'badge-finalizado'}
                st.markdown(f'<span class="{badge_map.get(r["status"], "badge-disponivel")}">{r["status"]}</span>',
                            unsafe_allow_html=True)
                st.markdown("---")

                col1, col2, col3, col4 = st.columns(4)
                col1.markdown(f"**Modelo Final**\n\n{r.get('modelo') or '—'}")
                col2.markdown(f"**Modelo Original**\n\n{r.get('modelo_original') or '—' if r.get('modelo_original') else '(sem transformação)'}")
                col3.markdown(f"**DI**\n\n{r.get('di') or '—'}")
                col4.markdown(f"**Nº Lacre**\n\n{r.get('numero_lacre') or '—'}")

                col5, col6 = st.columns(2)
                col5.markdown(f"**Cliente**\n\n{r.get('cliente') or '—'}")
                col6.markdown(f"**Pedido**\n\n{r.get('pedido') or '—'}")

                if any([r.get('valor_antes'), r.get('valor_depois'), r.get('exc_se'),
                        r.get('exc_sd'), r.get('exc_ie'), r.get('exc_id'),
                        r.get('carga_maxima'), r.get('zero')]):
                    st.markdown("---")
                    st.subheader("📐 Dados de Calibração")
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Valor Antes", r.get('valor_antes') or "—")
                    c2.metric("Valor Depois", r.get('valor_depois') or "—")
                    c3.metric("Carga Máx", r.get('carga_maxima') or "—")
                    st.markdown("**Excentricidade (4 cantos):**")
                    ce1, ce2, ce3, ce4 = st.columns(4)
                    ce1.metric("Sup. Esq", r.get('exc_se') or "—")
                    ce2.metric("Sup. Dir", r.get('exc_sd') or "—")
                    ce3.metric("Inf. Esq", r.get('exc_ie') or "—")
                    ce4.metric("Inf. Dir", r.get('exc_id') or "—")
                    st.metric("Zero", r.get('zero') or "—")

        else:
            # ── LISTA ─────────────────────────
            busca = st.text_input(
                "🔎 Buscar por Serial China, Serial Brasil, DI, Pedido, Cliente ou Modelo:")

            where = ""
            params = None
            if busca:
                where = """ WHERE di ILIKE %s OR serial_china ILIKE %s
                             OR serial_brasil ILIKE %s OR cliente ILIKE %s
                             OR pedido ILIKE %s OR modelo ILIKE %s"""
                p = f"%{busca}%"
                params = (p, p, p, p, p, p)

            df = query_df(f"""
                SELECT id, di AS "DI", modelo AS "Modelo",
                       modelo_original AS "Modelo Original",
                       cliente AS "Cliente", pedido AS "Pedido",
                       serial_china AS "Serial China",
                       serial_brasil AS "Serial Brasil",
                       numero_lacre AS "Nº Lacre",
                       status AS "Status"
                FROM producao {where} ORDER BY id DESC
            """, params=params)

            if not df.empty:
                st.markdown(f"**{len(df)} registro(s) encontrado(s)**")
                st.dataframe(df.drop(columns=['id']).style.map(color_status, subset=['Status']),
                             use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("**Ver detalhe de uma balança:**")
                opcoes_det = df.apply(
                    lambda r: f"{r['Serial Brasil'] or r['Serial China']}  |  {r['Modelo']}  |  {r['Status']}",
                    axis=1).tolist()

                sel = st.selectbox("Selecione", range(len(opcoes_det)),
                    format_func=lambda i: opcoes_det[i],
                    index=None, placeholder="Escolha para ver o detalhe...")

                if sel is not None:
                    st.session_state.detalhe_id = df['id'].tolist()[sel]
                    st.rerun()

                st.download_button("⬇️ Exportar para Excel",
                    data=gerar_excel(df.drop(columns=['id'])),
                    file_name="consulta_bel.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Nenhum registro encontrado.")

    # ══════════════════════════════════════════
    # CONFIGURAÇÕES
    # ══════════════════════════════════════════
    elif menu == "⚙️  Configurações":
        st.title("⚙️ Configurações do Sistema")
        st.markdown("---")
        st.subheader("⚠️ Zona de Manutenção")
        st.warning("Esta ação apaga **todos os dados** permanentemente e não pode ser desfeita.")
        if st.checkbox("Confirmo que desejo apagar todos os dados"):
            if st.button("🚨 RESETAR SISTEMA", type="primary"):
                conn = get_conn()
                c = conn.cursor()
                c.execute("DROP TABLE IF EXISTS modelos CASCADE")
                c.execute("DROP TABLE IF EXISTS clientes CASCADE")
                c.execute("DROP TABLE IF EXISTS producao CASCADE")
                conn.commit(); conn.close()
                init_db()
                st.success("Banco de dados recriado do zero!")
                st.rerun()

# ══════════════════════════════════════════════
# CHINA
# ══════════════════════════════════════════════
elif unidade == "China (Factory)":
    st.title("🇨🇳 China Production Interface")
    st.info("Módulo em desenvolvimento.")

# ══════════════════════════════════════════════
# TELA INICIAL
# ══════════════════════════════════════════════
else:
    st.sidebar.info("👆 Selecione uma unidade para começar.")
    st.title("🔬 Bel Engineering — Traceability System")
    st.markdown("""
    ### Bem-vindo ao Sistema de Rastreabilidade Global

    Selecione a unidade operacional no menu lateral para acessar as ferramentas.

    ---
    **Fluxo operacional:**
    1. **Gestão de Cadastros** — Cadastre modelos, clientes e registre a entrada das DIs
    2. **Bancada de Aferição** — Selecione as balanças, defina transformações, baixe a planilha e importe
    3. **Pronta Entrega** — Vincule clientes às balanças em estoque quando a venda acontecer
    4. **Consulta** — Pesquise qualquer balança, veja o detalhe completo e gere o relatório PDF
    5. **Dashboard** — Acompanhe o status geral do inventário
    """)
